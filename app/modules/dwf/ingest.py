from __future__ import annotations

import os
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.modules.dwf import models as dwf
from app.db import SessionLocal, engine, ensure_pg_extensions


def _clean_str(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() in ("nan", "none") else s


def _to_float(x: Any) -> Optional[float]:
    s = _clean_str(x)
    if not s or s == "-":
        return None
    try:
        return float(s)
    except Exception:
        return None


def _rows(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for raw in rows[1:]:
        if not any(raw):
            continue
        record = {headers[i]: raw[i] for i in range(len(headers))}
        data.append(record)
    return data


def _list_options(ws) -> List[dwf.DwfListOption]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean_str(h) for h in rows[0]]
    options: List[dwf.DwfListOption] = []
    for col_idx, list_ref in enumerate(headers):
        if not list_ref:
            continue
        values: List[str] = []
        for row in rows[1:]:
            if col_idx >= len(row):
                continue
            val = _clean_str(row[col_idx])
            if val:
                values.append(val)
        for idx, val in enumerate(values):
            options.append(
                dwf.DwfListOption(
                    list_ref=list_ref,
                    value=val,
                    display_order=idx,
                )
            )
    return options


def ingest(truncate_bank: bool = True) -> None:
    excel_file = os.getenv("DWF_EXCEL_FILE", "DWF_QuestionBank_Template.xlsx")
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"{excel_file} not found. Set DWF_EXCEL_FILE.")

    ensure_pg_extensions()
    dwf.Base.metadata.create_all(bind=engine)

    wb = load_workbook(excel_file)
    q_rows = _rows(wb["Questions"])
    a_rows = _rows(wb["AnswerOptions"])
    list_ws = wb["Lists"]

    db: Session = SessionLocal()
    try:
        if truncate_bank:
            db.execute(text("TRUNCATE TABLE dwf_answer_options RESTART IDENTITY CASCADE"))
            db.execute(text("TRUNCATE TABLE dwf_questions RESTART IDENTITY CASCADE"))
            db.execute(text("TRUNCATE TABLE dwf_list_options RESTART IDENTITY CASCADE"))
            db.flush()

        q_objects: List[dwf.DwfQuestion] = []
        for row in q_rows:
            qid = _clean_str(row.get("Q-ID"))
            if not qid:
                continue
            q_objects.append(
                dwf.DwfQuestion(
                    q_id=qid,
                    section=_clean_str(row.get("Section")) or None,
                    category=_clean_str(row.get("Category")) or None,
                    question_title=_clean_str(row.get("Question Title")) or None,
                    question_text=_clean_str(row.get("Question Text")),
                    guidance=_clean_str(row.get("Guidance")) or None,
                    input_type=_clean_str(row.get("InputType")) or "Single",
                    list_ref=_clean_str(row.get("ListRef")) or None,
                    metric_key=_clean_str(row.get("MetricKey")) or None,
                    weight=_to_float(row.get("Weight")),
                    persona_scope=_clean_str(row.get("PersonaScope")) or None,
                )
            )
        db.bulk_save_objects(q_objects)
        db.flush()

        a_objects: List[dwf.DwfAnswerOption] = []
        for row in a_rows:
            aid = _clean_str(row.get("A-ID"))
            qid = _clean_str(row.get("Q-ID"))
            if not aid or not qid:
                continue
            a_objects.append(
                dwf.DwfAnswerOption(
                    a_id=aid,
                    q_id=qid,
                    option_number=int(row["Option #"]) if _clean_str(row.get("Option #")) else None,
                    answer_text=_clean_str(row.get("Answer Option Text")),
                    base_score=float(_clean_str(row.get("BaseScore")) or 0.0),
                    tags=_clean_str(row.get("Tags")) or None,
                )
            )
        db.bulk_save_objects(a_objects)
        db.flush()

        list_objects = _list_options(list_ws)
        if list_objects:
            db.bulk_save_objects(list_objects)
            db.flush()

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    ingest(truncate_bank=bool(os.getenv("TRUNCATE_BANK", "").lower() in ("1", "true", "yes")))
